"""
core/auto_archive.py
-------------------
Background monitor that periodically checks all Excel sheets in the
bridge's data root and auto-archives sheets that exceed the configured
row-count threshold.

Archive naming scheme:
  First archive                → {sheetName}.old
  If {sheetName}.old exists    → {sheetName}.old1
  If {sheetName}.old1 exists   → {sheetName}.old2
  …and so on.

When a sheet is archived:
  - The original data is copied to the archive sheet (under a new name).
  - A note "new data on new sheet" is appended to the OLD (archived) sheet.
  - A note "old data on old sheet → {archiveName}" is prepended to the
    NEW (fresh) sheet.

The original sheet name always stays the same so that the frontend
continues writing to it uninterrupted.
"""

from __future__ import annotations

import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Excel absolute row limit — auto-archive regardless of user threshold
# ---------------------------------------------------------------------------
EXCEL_MAX_ROWS = 1_048_576


class AutoArchiveMonitor:
    """
    Background monitor that checks all Excel sheets at a configured
    interval and archives sheets whose row count exceeds the threshold.
    """

    def __init__(self, settings: Any) -> None:
        self._settings = settings
        self._thread: threading.Thread | None = None
        self._stop_event = threading.Event()
        self._lock = threading.Lock()

        # Runtime state (updated live so endpoints can report it)
        self.last_check_time: str | None = None
        self.next_check_time: str | None = None
        self.sheets_monitored: int = 0
        self.last_archive: dict[str, Any] | None = None  # result of most recent archive

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def start(self) -> None:
        """Launch the background monitoring thread (no-op if already running)."""
        with self._lock:
            if self._thread and self._thread.is_alive():
                return
            self._stop_event.clear()
            self._thread = threading.Thread(target=self._monitor_loop, daemon=True, name="auto-archive")
            self._thread.start()
            print("  🗄️  Auto-archive monitor started "
                  f"(interval={self._settings.AUTO_ARCHIVE_CHECK_INTERVAL_MINUTES}m, "
                  f"max_rows={self._settings.AUTO_ARCHIVE_MAX_ROWS})")

    def stop(self) -> None:
        """Signal the background thread to stop and wait for it."""
        self._stop_event.set()
        if self._thread and self._thread.is_alive():
            self._thread.join(timeout=10)

    def get_status(self) -> dict[str, Any]:
        """Return a snapshot of the monitor's current state."""
        return {
            "enabled": self._settings.AUTO_ARCHIVE_ENABLED,
            "max_rows": self._settings.AUTO_ARCHIVE_MAX_ROWS,
            "check_interval_minutes": self._settings.AUTO_ARCHIVE_CHECK_INTERVAL_MINUTES,
            "last_check": self.last_check_time,
            "next_check": self.next_check_time,
            "sheets_monitored": self.sheets_monitored,
            "last_archive": self.last_archive,
        }

    def check_now(self) -> dict[str, Any]:
        """Force an immediate check of all sheets. Returns what was done."""
        return self._check_all()

    def archive_sheet(self, workbook_path: str | Path, sheet_name: str) -> dict[str, Any]:
        """Manually archive a specific sheet. Returns the result."""
        path = Path(workbook_path).resolve()
        return self._archive_sheet(path, sheet_name)

    # ------------------------------------------------------------------
    # Internal — background loop
    # ------------------------------------------------------------------

    def _monitor_loop(self) -> None:
        """Main loop: sleep for the configured interval, then check all sheets."""
        while not self._stop_event.is_set():
            interval = max(self._settings.AUTO_ARCHIVE_CHECK_INTERVAL_MINUTES, 1)
            self.next_check_time = datetime.now(timezone.utc).isoformat()

            # Sleep in 5-second chunks so we can respond to stop quickly
            for _ in range(interval * 60 // 5):
                if self._stop_event.is_set():
                    return
                time.sleep(5)

            if self._stop_event.is_set():
                return

            if not self._settings.AUTO_ARCHIVE_ENABLED:
                self.last_check_time = None
                continue

            try:
                self._check_all()
            except Exception as exc:
                print(f"  ⚠️  Auto-archive check failed: {exc}")

    # ------------------------------------------------------------------
    # Internal — check & archive
    # ------------------------------------------------------------------

    def _check_all(self) -> dict[str, Any]:
        """
        Scan all .xlsx workbooks under DATA_ROOT and archive any sheet
        whose row count exceeds the threshold (or Excel absolute limit).
        """
        self.last_check_time = datetime.now(timezone.utc).isoformat()
        data_root = Path(self._settings.DATA_ROOT)
        archived: list[dict[str, Any]] = []
        total_sheets = 0

        if not data_root.is_dir():
            self.sheets_monitored = 0
            return {"checked_at": self.last_check_time, "archived": [], "total_sheets": 0}

        for xlsx_path in sorted(data_root.rglob("*.xlsx")):
            # Skip temp files
            if xlsx_path.suffix == ".tmp" or xlsx_path.name.startswith("~"):
                continue
            try:
                result = self._check_workbook(xlsx_path)
                total_sheets += result.get("sheets_checked", 0)
                archived.extend(result.get("archived", []))
            except Exception as exc:
                print(f"  ⚠️  Could not check {xlsx_path.name}: {exc}")

        self.sheets_monitored = total_sheets
        if archived:
            self.last_archive = {"time": self.last_check_time, "sheets": archived}
            print(f"  🗄️  Auto-archived {len(archived)} sheet(s) — {archived}")

        return {
            "checked_at": self.last_check_time,
            "archived": archived,
            "total_sheets": total_sheets,
        }

    def _check_workbook(self, path: Path) -> dict[str, Any]:
        """Check all sheets in a single workbook, archive if needed."""
        threshold = self._settings.AUTO_ARCHIVE_MAX_ROWS
        archived = []
        sheets_checked = 0

        # Quick check: skip if file is too small to have large sheets
        try:
            if path.stat().st_size < 1024:
                return {"sheets_checked": 0, "archived": []}
        except OSError:
            return {"sheets_checked": 0, "archived": []}

        # Determine sheets to check — read sheet names only (lightweight)
        try:
            xl = pd.ExcelFile(path)
            sheet_names = xl.sheet_names
        except Exception:
            return {"sheets_checked": 0, "archived": []}

        for sheet_name in sheet_names:
            # Skip archive sheets themselves (they have .old suffix)
            if ".old" in sheet_name.lower():
                continue

            sheets_checked += 1
            try:
                df = pd.read_excel(path, sheet_name=sheet_name)
                row_count = len(df)
            except Exception:
                continue

            # Archive if:
            #   1. Row count exceeds user threshold, OR
            #   2. Row count is within 10% of Excel's absolute limit
            if row_count >= threshold or row_count >= int(EXCEL_MAX_ROWS * 0.9):
                try:
                    result = self._archive_sheet(path, sheet_name)
                    archived.append(result)
                except Exception as exc:
                    print(f"  ⚠️  Failed to archive sheet '{sheet_name}' in {path.name}: {exc}")

        return {"sheets_checked": sheets_checked, "archived": archived}

    def _archive_sheet(self, path: Path, sheet_name: str) -> dict[str, Any]:
        """
        Archive a single sheet:
          1. Find the next available archive name ({sheetName}.old / .old1 / .old2 …)
          2. Copy the sheet to the archive name
          3. Clear the original sheet (keep header row)
          4. Add marker notes to both sheets
          5. Save the workbook
        """
        # ---- Find next archive name ----
        wb = openpyxl.load_workbook(path)
        archive_name = self._next_archive_name(wb.sheetnames, sheet_name)

        # ---- Get source sheet ----
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"sheet": sheet_name, "workbook": str(path), "error": "Sheet not found"}

        src_ws = wb[sheet_name]
        max_row = src_ws.max_row
        max_col = src_ws.max_column

        # ---- Copy to archive sheet ----
        dst_ws = wb.create_sheet(title=archive_name)
        for row in src_ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)

        # ---- Add note to OLD (archived) sheet ----
        note_row = max_row + 1
        dst_ws.cell(row=note_row, column=1, value="new data on new sheet")

        # ---- Clear original sheet (keep header) ----
        if max_row > 1:
            src_ws.delete_rows(2, max_row - 1)

        # ---- Add note to NEW (fresh) sheet ----
        src_ws.cell(row=2, column=1, value=f"old data on old sheet → {archive_name}")

        # ---- Save ----
        tmp_path = path.with_suffix(".tmp.xlsx")
        wb.save(tmp_path)
        wb.close()

        # Atomic replace
        import os as _os
        import shutil as _shutil
        try:
            _os.replace(tmp_path, path)
        except PermissionError:
            _shutil.copy2(tmp_path, path)
            tmp_path.unlink(missing_ok=True)

        return {
            "sheet": sheet_name,
            "archive_name": archive_name,
            "workbook": str(path),
            "rows_archived": max_row - 1,  # excluding header
            "archived_at": datetime.now(timezone.utc).isoformat(),
        }

    @staticmethod
    def _next_archive_name(existing_sheets: list[str], base_name: str) -> str:
        """
        Determine the next archive sheet name.

        Naming: {base}.old → {base}.old1 → {base}.old2 → ...
        """
        first = f"{base_name}.old"
        if first not in existing_sheets:
            return first

        counter = 1
        while True:
            candidate = f"{base_name}.old{counter}"
            if candidate not in existing_sheets:
                return candidate
            counter += 1
            if counter > 9999:
                # Should never happen in practice
                raise RuntimeError(f"Too many archive versions for sheet '{base_name}'")


# ---------------------------------------------------------------------------
# Singleton — created and managed by main.py
# ---------------------------------------------------------------------------
_monitor: AutoArchiveMonitor | None = None


def get_monitor() -> AutoArchiveMonitor | None:
    """Return the global AutoArchiveMonitor instance, or None if not started."""
    return _monitor


def start_monitor(settings: Any) -> AutoArchiveMonitor:
    """Create and start the global monitor. No-op if already running."""
    global _monitor
    if _monitor is not None and _monitor._thread and _monitor._thread.is_alive():
        return _monitor
    _monitor = AutoArchiveMonitor(settings)
    _monitor.start()
    return _monitor


def stop_monitor() -> None:
    """Stop the global monitor if running."""
    global _monitor
    if _monitor is not None:
        _monitor.stop()
        _monitor = None
