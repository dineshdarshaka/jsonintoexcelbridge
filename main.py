"""
main.py
-------
FastAPI entry-point for the Local Bridge Service.

Routes
------
GET  /data                    — Read the Excel file and return its contents as JSON.
POST /update                  — Decrypt an encrypted JSON payload and write it to Excel.
GET  /health                  — Simple health-check (no auth required).
GET  /admin                   — Visual configuration dashboard (HTML).
GET  /config                  — Get current config snapshot (secrets masked).
POST /config                  — Update config values and persist to .env.
GET  /config/generate-api-key — Generate a fresh API key.
GET  /config/generate-fernet  — Generate a fresh Fernet key.
"""

from __future__ import annotations

import argparse
import json
import os
import platform
import secrets as _secrets
import shutil
import socket
import sys
import time
from pathlib import Path
from typing import Any

import pandas as pd
import requests as _requests
from cryptography.fernet import Fernet
from fastapi import FastAPI, Request, HTTPException, status, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, HTMLResponse
from pydantic import BaseModel, Field

from core.config import settings, get_config_snapshot, update_env_file, detect_safe_path
from core.security import verify_api_key, decrypt_payload
from core.command_engine import CommandEngine

# ===================================================================
# FastAPI application
# ===================================================================

app = FastAPI(
    title="Local Bridge Service",
    description="Secure REST bridge between a web frontend and a local Excel file.",
    version="1.0.0",
)

# -------------------------------------------------------------------
# CORS — restrict to the web app's origin
# -------------------------------------------------------------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["X-API-KEY", "Content-Type"],
)

# ===================================================================
# Helper — safe atomic file replace on Windows
# ===================================================================

def _safe_atomic_replace(src: Path, dst: Path, retries: int = 3, delay: float = 0.5) -> None:
    """
    Atomically replace ``dst`` with ``src``.

    On Windows, ``os.replace`` fails with PermissionError if the target
    file is locked (e.g. open in Excel).  This helper retries several
    times, then falls back to a content-copy via ``shutil.copy2``.
    """
    if not src.is_file():
        raise FileNotFoundError(f"Source file not found: {src}")

    last_err: Exception | None = None

    for attempt in range(1, retries + 1):
        try:
            os.replace(src, dst)
            return  # success
        except PermissionError as exc:
            last_err = exc
            if attempt < retries:
                time.sleep(delay)
            continue

    # All retries exhausted — try copy-then-delete as last resort
    try:
        shutil.copy2(src, dst)
        src.unlink(missing_ok=True)
        return
    except Exception as exc:
        last_err = exc

    raise OSError(f"Cannot replace {dst}: {last_err}") from last_err


# ===================================================================
# Pydantic schemas
# ===================================================================


class UpdatePayload(BaseModel):
    """
    Expected POST body.

    The frontend sends an encrypted JSON blob.  The *encrypted* field
    carries the Fernet token (Base64-encoded).

    Optional sheet_name tells the bridge which Excel sheet to write to.
    If omitted, the configured default sheet is used.  If the sheet does
    not exist it will be created automatically.
    """

    encrypted: str = Field(
        ...,
        description="Fernet-encrypted, Base64-encoded JSON payload.",
        min_length=1,
    )
    sheet_name: str = Field(
        default="",
        description="Target Excel sheet name. Auto-created if missing.",
    )


class DataResponse(BaseModel):
    """Successful GET /data response shape."""

    status: str = "ok"
    sheet_name: str
    row_count: int
    columns: list[str]
    data: list[dict[str, Any]]


class UpdateResponse(BaseModel):
    """Successful POST /update response shape."""

    status: str = "ok"
    message: str
    rows_written: int


class ConfigUpdatePayload(BaseModel):
    """
    Payload for POST /config — a flat dict of key-value pairs.
    Only recognised keys are persisted; others are ignored.
    """

    updates: dict[str, str | int] = Field(
        ...,
        description="Key-value pairs to write to the .env file.",
        min_length=1,
    )


class ConfigSnapshot(BaseModel):
    """Response shape for GET /config."""

    config: dict[str, Any]


class ConfigUpdateResult(BaseModel):
    """Response shape for POST /config."""

    status: str
    results: dict[str, str]
    restart_required: bool = True


class CommandRequest(BaseModel):
    """
    Payload for POST /command — send a text command to the bridge.
    Commands: ADD, RMV, REPL, FIND, SHOW, HELP
    """

    command: str = Field(
        ...,
        description="Command string (e.g. 'ADD {\"col\":\"val\"}'). Prefix / optional.",
        min_length=1,
    )


# ===================================================================
# Dependency — every protected route runs this
# ===================================================================


async def require_auth(request: Request) -> None:
    """FastAPI dependency that enforces X-API-KEY validation."""
    verify_api_key(request)


async def require_main(request: Request) -> None:
    """FastAPI dependency that blocks writes if bridge is not main."""
    if not settings.BRIDGE_IS_MAIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="This bridge is not the main bridge for its location. Only the main bridge can save data to Excel.",
        )


# ===================================================================
# Approval callback (called by central app)
# ===================================================================

class ApprovalPayload(BaseModel):
    """Payload sent by the central app when approving/rejecting a bridge."""
    is_approved: bool = False
    is_main: bool = False
    location: str = ""
    office: str = ""


@app.post("/api/approval")
async def approval_callback(payload: ApprovalPayload, request: Request) -> dict:
    """
    Called by the central RMMS app when admin approves/rejects a bridge.
    Updates the local .env with approval status and main-bridge flag.
    """
    verify_api_key(request)
    updates = {
        "BRIDGE_IS_APPROVED": str(payload.is_approved).lower(),
        "BRIDGE_IS_MAIN": str(payload.is_main).lower(),
        "BRIDGE_LOCATION": payload.location,
        "BRIDGE_OFFICE": payload.office,
    }
    results = update_env_file(updates)
    # Reload settings so runtime sees the change
    settings.BRIDGE_IS_APPROVED = payload.is_approved
    settings.BRIDGE_IS_MAIN = payload.is_main
    settings.BRIDGE_LOCATION = payload.location
    settings.BRIDGE_OFFICE = payload.office
    return {"status": "ok", "updates": results}


@app.get("/health")
async def health() -> dict[str, str]:
    """Unprotected health-check endpoint."""
    return {"status": "healthy", "service": "local-bridge"}


@app.get("/", response_class=HTMLResponse)
async def root_page():
    """Bridge landing page — shows machine info + 'Send Report to Admin' form."""
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"

    return HTMLResponse(content=_ROOT_HTML.format(
        hostname=hostname,
        local_ip=local_ip,
        os_name=platform.system(),
        os_release=platform.release(),
        python_ver=platform.python_version(),
        port=settings.PORT,
        bridge_url=f"http://{local_ip}:{settings.PORT}",
        excel_file=str(settings.EXCEL_FILE_PATH),
        sheet_name=settings.EXCEL_SHEET_NAME,
        api_key=settings.API_KEY,
        fernet_key=settings.FERNET_KEY,
        bridge_id=settings.BRIDGE_ID,
        status_badge=(
            '<span style="color:#16a34a;">✅ Approved (Main)</span>' if settings.BRIDGE_IS_MAIN
            else '<span style="color:#16a34a;">✅ Approved</span>' if settings.BRIDGE_IS_APPROVED
            else '<span style="color:#92400e;">⏳ Pending Approval</span>'
        ),
        location_display=settings.BRIDGE_LOCATION or "<em style='color:#94a3b8;'>Not set</em>",
        loc_value=settings.BRIDGE_LOCATION or "",
    ))


@app.get("/info")
async def machine_info() -> dict[str, Any]:
    """
    Return auto-detected machine information.
    Useful for registering this bridge with the central web app.
    Secrets are shown as masked fingerprints — never exposed in full.
    """
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"

    return {
        "status": "ok",
        "bridge_id": settings.BRIDGE_ID,
        "machine": {
            "hostname": hostname,
            "local_ip": local_ip,
            "os": platform.system(),
            "os_version": platform.version(),
            "python": platform.python_version(),
            "bridge_port": settings.PORT,
            "bridge_url": f"http://{local_ip}:{settings.PORT}",
        },
        "config_summary": {
            "excel_file": str(settings.EXCEL_FILE_PATH),
            "sheet_name": settings.EXCEL_SHEET_NAME,
            "api_key_fingerprint": settings.API_KEY[:8] + "..." + settings.API_KEY[-4:],
            "fernet_key_fingerprint": settings.FERNET_KEY[:8] + "...",
            "allowed_origins": settings.ALLOWED_ORIGINS,
        },
    }


@app.get(
    "/data",
    response_model=DataResponse,
    dependencies=[Depends(require_auth)],
)
async def get_data(sheet: str = "") -> DataResponse:
    """
    Read the configured Excel file and return all rows as JSON.

    Query params:
      sheet  — Optional sheet name. If provided, reads from that sheet.
               Otherwise uses the configured default sheet.

    Returns 404 if the file or sheet does not exist.
    """
    path: Path = settings.EXCEL_FILE_PATH
    sheet_name = sheet.strip() if sheet else settings.EXCEL_SHEET_NAME

    if not path.is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Excel file not found at {path}",
        )

    try:
        # Check if the sheet exists
        xl = pd.ExcelFile(path)
        if sheet_name not in xl.sheet_names:
            # Sheet doesn't exist — return empty dataset (don't error)
            return DataResponse(
                sheet_name=sheet_name,
                row_count=0,
                columns=["Date & Time", "Barcode", "Action"],
                data=[],
            )
        df: pd.DataFrame = pd.read_excel(path, sheet_name=sheet_name)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Sheet '{sheet_name}' not found in {path}: {exc!s}",
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to read Excel file: {exc!s}",
        )

    # Replace NaN / NaT with None for clean JSON
    df = df.where(pd.notna(df), None)

    return DataResponse(
        sheet_name=sheet_name,
        row_count=len(df),
        columns=list(df.columns),
        data=df.to_dict(orient="records"),
    )


@app.post(
    "/update",
    response_model=UpdateResponse,
    dependencies=[Depends(require_auth), Depends(require_main)],
)
async def update_data(payload: UpdatePayload) -> UpdateResponse:
    """
    Decrypt the encrypted JSON payload and overwrite the Excel file
    with the provided data.

    The decrypted plaintext must be a JSON array of objects; each object
    becomes one row, and keys become column headers.

    Returns 400 on decryption / validation failures.
    """
    # 1. Decrypt
    plaintext: str = decrypt_payload(payload.encrypted)

    # 2. Parse the decrypted JSON
    try:
        records: list[dict[str, Any]] = json.loads(plaintext)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Decrypted payload is not valid JSON: {exc!s}",
        )

    if not isinstance(records, list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Decrypted payload must be a JSON array of objects.",
        )

    if not records:
        # Empty dataset = clear the sheet (keep the sheet, remove all rows)
        try:
            path: Path = settings.EXCEL_FILE_PATH
            target_sheet = payload.sheet_name.strip() if payload.sheet_name else settings.EXCEL_SHEET_NAME

            if path.is_file():
                import openpyxl
                tmp_path = path.with_suffix(".tmp.xlsx")
                wb = openpyxl.load_workbook(path)
                if target_sheet in wb.sheetnames:
                    ws = wb[target_sheet]
                    # Delete all rows except header
                    ws.delete_rows(1, ws.max_row)
                wb.save(tmp_path)
                wb.close()
                _safe_atomic_replace(tmp_path, path)
            return UpdateResponse(
                message=f"Sheet '{target_sheet}' cleared successfully.",
                rows_written=0,
            )
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to clear sheet: {exc!s}",
            )

    # 3. Convert to DataFrame and write
    try:
        df: pd.DataFrame = pd.DataFrame(records)
        path: Path = settings.EXCEL_FILE_PATH
        target_sheet = payload.sheet_name.strip() if payload.sheet_name else settings.EXCEL_SHEET_NAME

        # Ensure parent directory exists
        path.parent.mkdir(parents=True, exist_ok=True)

        # Write to a temporary file first, then atomically replace.
        tmp_path = path.with_suffix(".tmp.xlsx")

        # If the file already exists, we need to preserve other sheets.
        # Read existing workbook, update/add the target sheet, then write.
        if path.is_file():
            import openpyxl
            try:
                wb = openpyxl.load_workbook(path)
            except Exception:
                wb = openpyxl.Workbook()
            # Remove target sheet if it exists, then recreate it
            if target_sheet in wb.sheetnames:
                del wb[target_sheet]
            # If default sheet "Sheet" exists and we're writing to a different sheet, remove it
            if "Sheet" in wb.sheetnames and target_sheet != "Sheet" and len(wb.sheetnames) == 1:
                del wb["Sheet"]
            ws = wb.create_sheet(title=target_sheet)
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            # Write data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else None)
            wb.save(tmp_path)
            wb.close()
        else:
            df.to_excel(
                tmp_path,
                sheet_name=target_sheet,
                index=False,
            )
        _safe_atomic_replace(tmp_path, path)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to write Excel file: {exc!s}",
        )

    return UpdateResponse(
        message=f"Excel file updated successfully at {path}",
        rows_written=len(df),
    )


@app.post(
    "/command",
    dependencies=[Depends(require_auth)],
)
async def execute_command(payload: CommandRequest, request: Request):
    """
    Execute a command against the Excel database.

    Supported commands: ADD, RMV, REPL, FIND, SHOW, HELP

    Read-only commands (HELP, SHOW, GET, RANGE, LASTROW, LASTCOL, COLS) work
    on any bridge. Write commands require the bridge to be the main bridge.

    Examples:
      ADD  {"data": "new value"}
      RMV  data=hello
      REPL data=hello {"data": "updated value"}
      FIND data=hello
      SHOW
      HELP
    """
    engine = CommandEngine(
        excel_path=settings.EXCEL_FILE_PATH,
        sheet_name=settings.EXCEL_SHEET_NAME,
    )

    # Block write commands on non-main bridges
    if not CommandEngine.is_read_only(payload.command) and not settings.BRIDGE_IS_MAIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="This bridge is not the main bridge for its location. Only the main bridge can save data to Excel.",
        )

    result = engine.execute(payload.command)
    return result.to_dict()


# ===================================================================
# Configuration API (auth-protected)
# ===================================================================


@app.get(
    "/config",
    response_model=ConfigSnapshot,
    dependencies=[Depends(require_auth)],
)
async def get_config() -> ConfigSnapshot:
    """
    Return a snapshot of the current configuration.
    Secrets (API_KEY, FERNET_KEY) are **masked**.
    """
    return ConfigSnapshot(config=get_config_snapshot())


@app.post(
    "/config",
    response_model=ConfigUpdateResult,
    dependencies=[Depends(require_auth)],
)
async def update_config(payload: ConfigUpdatePayload) -> ConfigUpdateResult:
    """
    Update configuration values and persist them to the .env file.

    If the Excel file path is changed, the existing data file is
    **automatically moved** to the new location (preserving all sheets).

    The server must be **restarted** for server-level settings
    (HOST, PORT, CORS) to take effect.  API_KEY and FERNET_KEY take
    effect on the next request.
    """
    old_excel_path = settings.EXCEL_FILE_PATH
    new_excel_path_raw = payload.updates.get("EXCEL_FILE_PATH")

    results = update_env_file(payload.updates)

    # --- Auto-move Excel file if path changed ---
    if new_excel_path_raw:
        new_excel_path = Path(str(new_excel_path_raw).strip()).resolve()

        # Only move if the path actually changed and the old file exists
        if new_excel_path != old_excel_path.resolve() and old_excel_path.is_file():
            try:
                # Ensure the target directory exists
                new_excel_path.parent.mkdir(parents=True, exist_ok=True)

                # Copy the old file to the new location
                shutil.copy2(str(old_excel_path), str(new_excel_path))
                results["EXCEL_FILE_MOVED"] = (
                    f"Data copied from '{old_excel_path}' → '{new_excel_path}'."
                    f" You may delete the old file manually if desired."
                )
            except Exception as exc:
                results["EXCEL_FILE_MOVED"] = (
                    f"WARNING: Could not copy data to new location: {exc}. "
                    f"Old file still at '{old_excel_path}'. "
                    f"Please manually copy it to '{new_excel_path}'."
                )

    return ConfigUpdateResult(
        status="ok",
        results=results,
        restart_required=True,
    )


@app.get(
    "/config/generate-api-key",
    dependencies=[Depends(require_auth)],
)
async def generate_api_key() -> dict[str, str]:
    """Generate a cryptographically random API key (hex string)."""
    return {"api_key": _secrets.token_hex(32)}


@app.get(
    "/config/generate-fernet",
    dependencies=[Depends(require_auth)],
)
async def generate_fernet_key() -> dict[str, str]:
    """Generate a fresh Fernet encryption key."""
    return {"fernet_key": Fernet.generate_key().decode("utf-8")}


@app.get(
    "/config/safe-path",
    dependencies=[Depends(require_auth)],
)
async def get_safe_path() -> dict[str, str]:
    """
    Auto-detect the best non-system partition for Excel data storage.
    Returns the recommended path and a human-readable reason.
    """
    return detect_safe_path()


# ===================================================================
# Admin dashboard (HTML page)
# ===================================================================

# The admin page is served as a standalone HTML document.  The frontend
# validates the user's API key via a simple login form and then loads
# the full configuration panel.

_ROADBLOCK_GUARD = """
<script>
(function(){{
    if (window.top !== window.self) {{
        document.body.innerHTML = '<h2 style="color:red;text-align:center;margin-top:80px;">'
          + '⛔ This page cannot be embedded in an iframe.</h2>';
        document.body.style.display = 'block';
        throw new Error('Embedding not allowed');
    }}
}})();
</script>
"""

_ROOT_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Data Bridge — {hostname}</title>
{roadblock}
<style>
    :root {{
        --bg: #f8fafc;
        --card-bg: #ffffff;
        --text: #0f172a;
        --text-secondary: #475569;
        --primary: #2563eb;
        --success: #16a34a;
        --danger: #dc2626;
        --warning: #d97706;
        --border: #e2e8f0;
        --shadow: 0 4px 16px rgba(0,0,0,0.06);
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
        font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
        background: var(--bg);
        color: var(--text);
        min-height: 100vh;
        display: flex; flex-direction: column;
    }}
    .container {{ max-width: 700px; margin: 0 auto; padding: 2rem 1.5rem; width: 100%; }}
    .hero {{
        background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
        color: #fff; padding: 2.5rem 2rem;
        border-radius: 0 0 24px 24px;
        text-align: center;
    }}
    .hero h1 {{ font-size: 1.6rem; font-weight: 700; }}
    .hero p {{ opacity: 0.85; margin-top: 0.4rem; font-size: 0.95rem; }}
    .card {{
        background: var(--card-bg); border: 1px solid var(--border);
        border-radius: 16px; padding: 1.5rem; margin-top: 1.5rem;
        box-shadow: var(--shadow);
    }}
    .card h2 {{
        font-size: 1.1rem; margin-bottom: 1rem;
        display: flex; align-items: center; gap: 0.4rem;
    }}
    .info-grid {{
        display: grid; grid-template-columns: 1fr 1fr;
        gap: 0.75rem;
    }}
    @media (max-width: 500px) {{ .info-grid {{ grid-template-columns: 1fr; }} }}
    .info-item {{
        background: var(--bg); border-radius: 10px; padding: 0.75rem 1rem;
    }}
    .info-item .label {{
        font-size: 0.7rem; text-transform: uppercase;
        letter-spacing: 0.05em; color: var(--text-secondary);
        font-weight: 600;
    }}
    .info-item .value {{
        font-size: 0.9rem; font-weight: 600; margin-top: 0.2rem;
        word-break: break-all;
    }}
    input, button, select {{
        font-family: inherit; font-size: 0.95rem;
    }}
    input {{
        width: 100%; padding: 0.65rem 0.85rem;
        border: 1px solid var(--border); border-radius: 10px;
        background: var(--bg); color: var(--text);
        margin-bottom: 0.75rem;
    }}
    input:focus {{ outline: 2px solid var(--primary); outline-offset: -1px; }}
    label {{
        display: block; margin-bottom: 0.2rem;
        font-weight: 600; font-size: 0.85rem;
    }}
    .btn {{
        display: inline-block; padding: 0.6rem 1.4rem;
        border-radius: 10px; border: none; cursor: pointer;
        font-weight: 700; font-size: 0.95rem;
        transition: all 0.2s; text-decoration: none;
    }}
    .btn:hover {{ transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,0.15); }}
    .btn:active {{ transform: translateY(0); }}
    .btn-primary {{ background: var(--primary); color: #fff; }}
    .btn-success {{ background: var(--success); color: #fff; }}
    .btn-disabled {{ background: #94a3b8; color: #fff; cursor: not-allowed; }}
    .btn-disabled:hover {{ transform: none; box-shadow: none; }}
    .status {{ margin-top: 1rem; padding: 0.75rem 1rem; border-radius: 10px; font-weight: 600; font-size: 0.9rem; }}
    .status-ok {{ background: #dcfce7; color: #166534; }}
    .status-err {{ background: #fee2e2; color: #991b1b; }}
    .status-pending {{ background: #fef3c7; color: #92400e; }}
    .spinner {{
        display: inline-block; width: 16px; height: 16px;
        border: 2px solid #fff; border-top-color: transparent;
        border-radius: 50%; animation: spin 0.6s linear infinite;
        vertical-align: middle; margin-right: 0.3rem;
    }}
    @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
    .footer {{
        text-align: center; margin-top: 2rem;
        color: var(--text-secondary); font-size: 0.8rem;
    }}
</style>
</head>
<body>
<div class="hero">
    <h1>🖥️ Data Bridge</h1>
    <p id="heroHost">{hostname}</p>
</div>

<div class="container">

    <!-- Machine Info -->
    <div class="card">
        <h2>📊 Machine Information</h2>
        <div class="info-grid">
            <div class="info-item">
                <div class="label">Hostname</div>
                <div class="value">{hostname}</div>
            </div>
            <div class="info-item">
                <div class="label">Local IP</div>
                <div class="value">{local_ip}</div>
            </div>
            <div class="info-item">
                <div class="label">Bridge URL</div>
                <div class="value">{bridge_url}</div>
            </div>
            <div class="info-item">
                <div class="label">OS</div>
                <div class="value">{os_name} {os_release}</div>
            </div>
            <div class="info-item">
                <div class="label">Python</div>
                <div class="value">{python_ver}</div>
            </div>
            <div class="info-item">
                <div class="label">Excel File</div>
                <div class="value">{excel_file}</div>
            </div>
            <div class="info-item">
                <div class="label">Bridge ID</div>
                <div class="value" style="font-family:monospace; font-size:0.8rem; word-break:break-all;">{bridge_id}</div>
            </div>
            <div class="info-item">
                <div class="label">Status</div>
                <div class="value">{status_badge}</div>
            </div>
            <div class="info-item">
                <div class="label">Location</div>
                <div class="value">{location_display}</div>
            </div>
        </div>
    </div>

    <!-- Registration Request -->
    <div class="card">
        <h2>📩 Send Registration Request to Admin</h2>
        <p style="color:var(--text-secondary); font-size:0.85rem; margin-bottom:1rem;">
            Give this bridge a friendly name and fill in the Central App URL.
            The admin will review and approve your request.
        </p>

        <label for="bridgeName">Bridge Name <span style="color:var(--danger);">*</span></label>
        <input type="text" id="bridgeName"
               placeholder="e.g. Colombo Office"
               value="{hostname}">

        <label for="bridgeLoc">Location <span style="color:var(--danger);">*</span></label>
        <input type="text" id="bridgeLoc"
               list="locDatalist"
               placeholder="e.g. Colombo, Sri Lanka"
               value="{loc_value}" required>
        <datalist id="locDatalist"></datalist>

        <label for="bridgeOffice">Office <span style="color:var(--danger);">*</span></label>
        <input type="text" id="bridgeOffice"
               placeholder="e.g. Colombo Main, Nugegoda Branch"
               value="">
        <p style="color:var(--text-secondary); font-size:0.75rem; margin-top:-0.25rem; margin-bottom:0.75rem;">
            Office within the selected location (e.g. branch name, division, sub-area).
        </p>

        <label for="centralUrl">Central App URL <span style="color:var(--danger);">*</span></label>
        <input type="url" id="centralUrl"
               placeholder="http://192.168.1.10:5001"
               value="" onchange="fetchLocations()" onfocus="fetchLocations()">

        <div style="padding:0.5rem 0;">
            <label style="display:flex; align-items:center; gap:0.5rem; cursor:pointer; font-weight:400;">
                <input type="checkbox" id="bridgeIsMain" value="1" style="width:auto; margin:0; accent-color:var(--primary);">
                ⭐ Set as Main Bridge for this location (only one main per location)
            </label>
        </div>

        <button class="btn btn-primary" id="sendBtn" onclick="sendRequest()">
            📩 Send Request to Admin
        </button>

        <div id="statusArea" style="display:none;"></div>
    </div>

    <!-- Command Console -->
    <div class="card">
        <h2>💻 Command Console</h2>
        <p style="color:var(--text-secondary); font-size:0.85rem; margin-bottom:0.5rem;">
            Send commands to manage the Excel database. Type <code>HELP</code> for available commands.
        </p>

        <div style="display:flex; gap:0.5rem; align-items:flex-start;">
            <input type="text" id="cmdInput"
                   placeholder="e.g. ADD {{&quot;data&quot;:&quot;hello&quot;}} or RMV data=hello or SHOW"
                   style="flex:1; font-family:monospace; font-size:0.85rem;"
                   onkeydown="if(event.key==='Enter')sendCmd()">
            <button class="btn btn-primary" id="cmdBtn" onclick="sendCmd()" style="white-space:nowrap;">
                ▶ Run
            </button>
        </div>

        <div id="cmdStatus" style="display:none; margin-top:0.75rem;"></div>

        <details style="margin-top:0.75rem;">
            <summary style="cursor:pointer; color:var(--primary); font-weight:600; font-size:0.85rem;">
                📋 Output
            </summary>
            <pre id="cmdOutput" style="
                background:#1e293b; color:#e2e8f0; border-radius:8px;
                padding:0.75rem; margin-top:0.5rem; font-size:0.8rem;
                max-height:300px; overflow-y:auto; white-space:pre-wrap;
                word-break:break-word;
            "></pre>
        </details>
    </div>

    <div class="footer">
        Data Convert Engine &bull; Port {port}
    </div>
</div>

<script>
var submitting = false;
var locationsFetched = false;
var cmdRunning = false;
async function sendCmd() {{
    if (cmdRunning) return;
    var input = document.getElementById('cmdInput');
    var cmd = input.value.trim();
    if (!cmd) return;
    var btn = document.getElementById('cmdBtn');
    var statusEl = document.getElementById('cmdStatus');
    var outputEl = document.getElementById('cmdOutput');
    cmdRunning = true;
    btn.disabled = true;
    btn.textContent = '...';
    statusEl.style.display = 'block';
    statusEl.textContent = '⏳ Running...';
    statusEl.className = 'status status-pending';
    try {{
        var resp = await fetch('/command', {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json', 'X-API-KEY': '{api_key}' }},
            body: JSON.stringify({{ command: cmd }})
        }});
        var data = await resp.json();
        outputEl.textContent = JSON.stringify(data, null, 2);
        if (data.success) {{
            statusEl.textContent = '✅ ' + (data.message || 'Done.');
            statusEl.className = 'status status-ok';
        }} else {{
            statusEl.textContent = '❌ ' + (data.message || 'Failed.');
            statusEl.className = 'status status-err';
        }}
    }} catch(e) {{
        outputEl.textContent = 'Error: ' + e.message;
        statusEl.textContent = '❌ Could not reach bridge.';
        statusEl.className = 'status status-err';
    }}
    cmdRunning = false;
    btn.disabled = false;
    btn.textContent = '▶ Run';
}}
async function fetchLocations() {{
    var centralUrl = document.getElementById('centralUrl').value.trim();
    if (!centralUrl) return;
    if (locationsFetched) return;
    locationsFetched = true;
    try {{
        var resp = await fetch(centralUrl.replace(/\\/$/, '') + '/api/locations', {{ timeout: 5000 }});
        if (!resp.ok) throw new Error('non-ok');
        var locations = await resp.json();
        var dl = document.getElementById('locDatalist');
        dl.innerHTML = '';
        locations.forEach(function(loc) {{
            var opt = document.createElement('option');
            opt.value = loc;
            dl.appendChild(opt);
        }});
    }} catch(e) {{
        locationsFetched = false;  // allow retry if failed
    }}
}}
function showStatus(msg, type) {{
    var el = document.getElementById('statusArea');
    el.style.display = 'block';
    el.textContent = msg;
    el.className = 'status status-' + type;
}}
async function sendRequest() {{
    if (submitting) return;
    var bridgeName = document.getElementById('bridgeName').value.trim();
    var bridgeLoc = document.getElementById('bridgeLoc').value.trim();
    var bridgeOffice = document.getElementById('bridgeOffice').value.trim();
    var centralUrl = document.getElementById('centralUrl').value.trim();
    if (!bridgeName) {{
        showStatus('Please enter a bridge name.', 'err');
        return;
    }}
    if (!bridgeLoc) {{
        showStatus('Please enter the bridge location.', 'err');
        return;
    }}
    if (!bridgeOffice) {{
        showStatus('Please enter the office name.', 'err');
        return;
    }}
    if (!centralUrl) {{
        showStatus('Please enter the Central App URL.', 'err');
        return;
    }}
    var btn = document.getElementById('sendBtn');
    submitting = true;
    btn.disabled = true;
    btn.innerHTML = '<span class="spinner"></span> Sending...';
    btn.className = 'btn btn-disabled';
    showStatus('Submitting registration request...', 'pending');

    var body = {{
        id: "{bridge_id}",
        name: bridgeName,
        location: bridgeLoc,
        office: bridgeOffice,
        url: "{bridge_url}",
        api_key: "{api_key}",
        fernet_key: "{fernet_key}",
        is_main: document.getElementById('bridgeIsMain').checked
    }};

    try {{
        var resp = await fetch(centralUrl.replace(/\\/$/, '') + '/api/submit-registration', {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify(body)
        }});
        var data = await resp.json();
        if (resp.ok) {{
            var msg = '✅ ' + (data.message || 'Request submitted! Waiting for admin approval.');
            if (data.warning) msg += '\\n⚠️ ' + data.warning;
            if (data.main_status) msg += '\\n⭐ ' + data.main_status;
            showStatus(msg, 'ok');
            btn.innerHTML = '✅ Request Sent';
            btn.className = 'btn btn-success';
        }} else {{
            showStatus('❌ ' + (data.message || 'Server rejected the request.'), 'err');
            btn.innerHTML = '📩 Send Request to Admin';
            btn.className = 'btn btn-primary';
            btn.disabled = false;
            submitting = false;
        }}
    }} catch(e) {{
        showStatus('❌ Could not reach the central app at ' + centralUrl + '. Check the URL and make sure the app is running.', 'err');
        btn.innerHTML = '📩 Send Request to Admin';
        btn.className = 'btn btn-primary';
        btn.disabled = false;
        submitting = false;
    }}
}}
</script>
</body>
</html>
""".replace("{roadblock}", _ROADBLOCK_GUARD)

_ADMIN_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Local Bridge — Admin</title>
<style>
  :root {
    --bg: #f4f6f9;
    --card-bg: #ffffff;
    --text: #1a1a2e;
    --sub: #6b7280;
    --accent: #4f46e5;
    --accent-hover: #4338ca;
    --danger: #dc2626;
    --success: #16a34a;
    --border: #e5e7eb;
    --input-bg: #f9fafb;
    --shadow: 0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
    --radius: 10px;
  }
  @media (prefers-color-scheme: dark) {
    :root {
      --bg: #0f172a;
      --card-bg: #1e293b;
      --text: #f1f5f9;
      --sub: #94a3b8;
      --accent: #818cf8;
      --accent-hover: #6366f1;
      --border: #334155;
      --input-bg: #0f172a;
      --shadow: 0 1px 3px rgba(0,0,0,.4);
    }
  }
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Inter', system-ui, -apple-system, sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 20px;
  }
  .container {
    width: 100%;
    max-width: 680px;
  }
  .card {
    background: var(--card-bg);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    padding: 32px;
  }
  .card-header {
    margin-bottom: 24px;
  }
  .card-header h1 {
    font-size: 1.5rem;
    font-weight: 700;
  }
  .card-header p {
    color: var(--sub);
    font-size: .875rem;
    margin-top: 4px;
  }
  .form-group {
    margin-bottom: 18px;
  }
  .form-group label {
    display: block;
    font-size: .8125rem;
    font-weight: 600;
    margin-bottom: 5px;
    color: var(--sub);
    text-transform: uppercase;
    letter-spacing: .04em;
  }
  .input-row {
    display: flex;
    gap: 8px;
  }
  .input-row input {
    flex: 1;
  }
  input, textarea {
    width: 100%;
    padding: 10px 12px;
    font-size: .875rem;
    font-family: 'JetBrains Mono', 'Fira Code', monospace;
    border: 1px solid var(--border);
    border-radius: 6px;
    background: var(--input-bg);
    color: var(--text);
    outline: none;
    transition: border-color .15s;
  }
  input:focus, textarea:focus {
    border-color: var(--accent);
    box-shadow: 0 0 0 3px rgba(79,70,229,.15);
  }
  input.secret { -webkit-text-security: disc; }
  .btn {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 8px 16px;
    font-size: .8125rem;
    font-weight: 600;
    border: none;
    border-radius: 6px;
    cursor: pointer;
    transition: background .15s, transform .1s;
    white-space: nowrap;
  }
  .btn:active { transform: scale(.97); }
  .btn-primary { background: var(--accent); color: #fff; }
  .btn-primary:hover { background: var(--accent-hover); }
  .btn-outline {
    background: transparent;
    border: 1px solid var(--border);
    color: var(--text);
  }
  .btn-outline:hover { background: var(--border); }
  .btn-success { background: var(--success); color: #fff; }
  .btn-danger { background: var(--danger); color: #fff; }
  .btn-sm { padding: 5px 10px; font-size: .75rem; }
  .btn-block { width: 100%; justify-content: center; padding: 12px; font-size: .875rem; }
  .actions {
    display: flex;
    gap: 10px;
    margin-top: 24px;
  }
  .toast {
    padding: 12px 16px;
    border-radius: 6px;
    font-size: .8125rem;
    font-weight: 500;
    margin-bottom: 16px;
    display: none;
  }
  .toast.show { display: block; }
  .toast-error { background: #fef2f2; color: var(--danger); border: 1px solid #fecaca; }
  .toast-success { background: #f0fdf4; color: var(--success); border: 1px solid #bbf7d0; }
  .badge {
    display: inline-block;
    font-size: .6875rem;
    font-weight: 600;
    padding: 2px 8px;
    border-radius: 999px;
    margin-left: 6px;
  }
  .badge-warn { background: #fef3c7; color: #b45309; }
  .divider {
    border: none;
    border-top: 1px solid var(--border);
    margin: 20px 0;
  }
  .hidden { display: none !important; }
  .secret-row {
    display: flex;
    align-items: center;
    gap: 6px;
    margin-top: 4px;
  }
  .secret-row .toggle-vis {
    background: none;
    border: none;
    cursor: pointer;
    font-size: .75rem;
    color: var(--accent);
    white-space: nowrap;
  }
  .spinner {
    display: inline-block;
    width: 14px; height: 14px;
    border: 2px solid var(--border);
    border-top-color: var(--accent);
    border-radius: 50%;
    animation: spin .6s linear infinite;
    margin-right: 6px;
  }
  @keyframes spin { to { transform: rotate(360deg); } }
</style>
</head>
<body>
<div class="container">

  <!-- ========== LOGIN SCREEN ========== -->
  <div id="login-screen" class="card">
    <div class="card-header">
      <h1>&#128274; Local Bridge Admin</h1>
      <p>Enter the current API key from your <code>.env</code> file.</p>
    </div>
    <div id="login-toast" class="toast toast-error"></div>
    <div class="form-group">
      <label for="login-key">API Key</label>
      <input id="login-key" type="password" placeholder="Paste your API key..."
             autocomplete="off" spellcheck="false">
    </div>
    <button class="btn btn-primary btn-block" onclick="doLogin()">
      Unlock Dashboard
    </button>
  </div>

  <!-- ========== MAIN DASHBOARD ========== -->
  <div id="dashboard-screen" class="card hidden">

    <!-- Header -->
    <div class="card-header" style="display:flex;justify-content:space-between;align-items:center;">
      <div>
        <h1>&#9881;&#65039; Configuration Dashboard</h1>
        <p>Manage all Local Bridge settings.</p>
      </div>
      <button class="btn btn-outline btn-sm" onclick="doLogout()">Lock</button>
    </div>

    <!-- Toast -->
    <div id="toast" class="toast"></div>

    <!-- Form -->
    <form id="config-form" onsubmit="return false;">

      <!-- API Key -->
      <div class="form-group">
        <label for="cfg-api-key">API Key <span class="badge badge-warn">Secret</span></label>
        <div class="secret-row">
          <input type="password" id="cfg-api-key"
                 style="flex:1" placeholder="(masked)" autocomplete="off" spellcheck="false">
          <button type="button" class="btn btn-outline btn-sm toggle-vis"
                  onclick="toggleVisibility('cfg-api-key', this)">Show</button>
          <button type="button" class="btn btn-outline btn-sm"
                  onclick="generateAndSet('api-key')">&#127922; Generate</button>
        </div>
      </div>

      <!-- Fernet Key -->
      <div class="form-group">
        <label for="cfg-fernet-key">Fernet Key <span class="badge badge-warn">Secret</span></label>
        <div class="secret-row">
          <input type="password" id="cfg-fernet-key"
                 style="flex:1" placeholder="(masked)" autocomplete="off" spellcheck="false">
          <button type="button" class="btn btn-outline btn-sm toggle-vis"
                  onclick="toggleVisibility('cfg-fernet-key', this)">Show</button>
          <button type="button" class="btn btn-outline btn-sm"
                  onclick="generateAndSet('fernet')">&#127922; Generate</button>
        </div>
      </div>

      <hr class="divider">

      <!-- Allowed Origins -->
      <div class="form-group">
        <label for="cfg-origins">Allowed Origins (comma-separated)</label>
        <input type="text" id="cfg-origins" placeholder="http://localhost:3000">
      </div>

      <!-- Excel File Path -->
      <div class="form-group">
        <label for="cfg-excel-path">Excel File Path</label>
        <div class="input-row">
          <input type="text" id="cfg-excel-path" placeholder="data.xlsx">
          <button type="button" class="btn btn-outline btn-sm"
                  onclick="detectSafePath()" title="Auto-detect safe partition for data">&#128194; Safe Path</button>
        </div>
        <p style="font-size:.7rem;color:var(--sub);margin-top:4px;" id="excel-path-hint">
          Store data on a non-system drive to keep it safe if Windows is reformatted.
        </p>
      </div>

      <!-- Sheet Name -->
      <div class="form-group">
        <label for="cfg-sheet">Sheet Name</label>
        <input type="text" id="cfg-sheet" placeholder="Sheet1">
      </div>

      <hr class="divider">

      <!-- Host / Port side by side -->
      <div style="display:flex;gap:12px;">
        <div class="form-group" style="flex:2;">
          <label for="cfg-host">Host</label>
          <input type="text" id="cfg-host" placeholder="127.0.0.1">
        </div>
        <div class="form-group" style="flex:1;">
          <label for="cfg-port">Port</label>
          <input type="number" id="cfg-port" placeholder="8000">
        </div>
      </div>

      <!-- Actions -->
      <div class="actions">
        <button type="button" class="btn btn-primary btn-block" onclick="saveConfig()">
          Save Configuration
        </button>
      </div>
      <p style="font-size:.75rem;color:var(--sub);margin-top:10px;text-align:center;">
        &#9888;&#65039; After saving, <strong>restart the server</strong> for HOST/PORT/CORS changes to apply.
      </p>
    </form>
  </div>

</div>

<script>
// ------------------------------------------------------------------
// State
// ------------------------------------------------------------------
let API_KEY = '';

function headers() {
  return {
    'Content-Type': 'application/json',
    'X-API-KEY': API_KEY,
  };
}

// ------------------------------------------------------------------
// Toast
// ------------------------------------------------------------------
function showToast(msg, type) {
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.className = 'toast toast-' + type + ' show';
  setTimeout(function(){ t.className = 'toast'; }, 5000);
}

function showLoginToast(msg) {
  const t = document.getElementById('login-toast');
  t.textContent = msg;
  t.className = 'toast toast-error show';
}

// ------------------------------------------------------------------
// Login / Logout
// ------------------------------------------------------------------
async function doLogin() {
  const key = document.getElementById('login-key').value.trim();
  if (!key) { showLoginToast('Please enter your API key.'); return; }
  API_KEY = key;
  // Verify the key by fetching config
  try {
    const res = await fetch('/config', { headers: headers() });
    if (!res.ok) {
      if (res.status === 403) { showLoginToast('Invalid API key — access denied.'); }
      else { showLoginToast('Server error: ' + res.status); }
      API_KEY = '';
      return;
    }
    const data = await res.json();
    populateForm(data.config);
    document.getElementById('login-screen').classList.add('hidden');
    document.getElementById('dashboard-screen').classList.remove('hidden');
  } catch (e) {
    showLoginToast('Cannot reach the server. Is it running?');
    API_KEY = '';
  }
}

function doLogout() {
  API_KEY = '';
  document.getElementById('login-key').value = '';
  document.getElementById('dashboard-screen').classList.add('hidden');
  document.getElementById('login-screen').classList.remove('hidden');
}

// Allow Enter key on login
document.getElementById('login-key').addEventListener('keydown', function(e) {
  if (e.key === 'Enter') doLogin();
});

// ------------------------------------------------------------------
// Populate form from GET /config response
// ------------------------------------------------------------------
function populateForm(cfg) {
  // Secrets are already masked by the server — show placeholders
  document.getElementById('cfg-api-key').value = cfg.API_KEY || '';
  document.getElementById('cfg-fernet-key').value = cfg.FERNET_KEY || '';
  document.getElementById('cfg-origins').value = cfg.ALLOWED_ORIGINS || '';
  document.getElementById('cfg-excel-path').value = cfg.EXCEL_FILE_PATH || '';
  document.getElementById('cfg-sheet').value = cfg.EXCEL_SHEET_NAME || '';
  document.getElementById('cfg-host').value = cfg.HOST || '';
  document.getElementById('cfg-port').value = cfg.PORT || '';
}

// ------------------------------------------------------------------
// Toggle visibility
// ------------------------------------------------------------------
function toggleVisibility(inputId, btn) {
  const inp = document.getElementById(inputId);
  if (inp.type === 'password') { inp.type = 'text'; btn.textContent = 'Hide'; }
  else { inp.type = 'password'; btn.textContent = 'Show'; }
}

// ------------------------------------------------------------------
// Generate keys
// ------------------------------------------------------------------
async function generateAndSet(type) {
  const endpoint = type === 'api-key' ? '/config/generate-api-key' : '/config/generate-fernet';
  const field = type === 'api-key' ? 'cfg-api-key' : 'cfg-fernet-key';
  const keyField = type === 'api-key' ? 'api_key' : 'fernet_key';
  try {
    const res = await fetch(endpoint, { headers: headers() });
    if (!res.ok) throw new Error('Generation failed (HTTP ' + res.status + ')');
    const data = await res.json();
    const inp = document.getElementById(field);
    inp.value = data[keyField];
    inp.type = 'text'; // show the freshly generated key
    const toggleBtn = inp.parentElement.querySelector('.toggle-vis');
    if (toggleBtn) toggleBtn.textContent = 'Hide';
  } catch (e) {
    showToast('Failed to generate key: ' + e.message, 'error');
  }
}

// ------------------------------------------------------------------
// Detect safe Excel path
// ------------------------------------------------------------------
async function detectSafePath() {
  try {
    const res = await fetch('/config/safe-path', { headers: headers() });
    if (!res.ok) throw new Error((await res.json()).detail || 'Detection failed');
    const data = await res.json();
    document.getElementById('cfg-excel-path').value = data.path;
    document.getElementById('excel-path-hint').textContent =
      '✅ ' + data.reason;
    showToast('Safe path detected! Click Save to apply.', 'success');
  } catch (e) {
    showToast('Detection failed: ' + e.message, 'error');
  }
}

// ------------------------------------------------------------------
// Save config
// ------------------------------------------------------------------
async function saveConfig() {
  const updates = {
    'API_KEY':        document.getElementById('cfg-api-key').value.trim(),
    'FERNET_KEY':     document.getElementById('cfg-fernet-key').value.trim(),
    'ALLOWED_ORIGINS': document.getElementById('cfg-origins').value.trim(),
    'EXCEL_FILE_PATH': document.getElementById('cfg-excel-path').value.trim(),
    'EXCEL_SHEET_NAME': document.getElementById('cfg-sheet').value.trim(),
    'HOST':           document.getElementById('cfg-host').value.trim(),
    'PORT':           parseInt(document.getElementById('cfg-port').value.trim(), 10),
  };

  // Filter out empty / masked-only values (user didn't change them)
  const cleaned = {};
  for (const [k, v] of Object.entries(updates)) {
    if (v === '' || (typeof v === 'number' && isNaN(v))) continue;
    // Skip masked placeholders (the server returns masked values)
    if ((k === 'API_KEY' || k === 'FERNET_KEY') && String(v).includes('****')) continue;
    cleaned[k] = v;
  }

  if (Object.keys(cleaned).length === 0) {
    showToast('No changes to save.', 'error');
    return;
  }

  try {
    const res = await fetch('/config', {
      method: 'POST',
      headers: headers(),
      body: JSON.stringify({ updates: cleaned }),
    });
    const data = await res.json();
    if (!res.ok) throw new Error(data.detail || 'Save failed');

    let msg = 'Configuration saved! Restart the server for all changes to take effect.';
    if (data.results && data.results.EXCEL_FILE_MOVED) {
      msg = '✅ ' + data.results.EXCEL_FILE_MOVED.replace(/\\n/g, ' ');
    }
    showToast(msg, 'success');
    // Refresh the displayed config
    const cfgRes = await fetch('/config', { headers: headers() });
    if (cfgRes.ok) {
      const cfgData = await cfgRes.json();
      populateForm(cfgData.config);
    }
  } catch (e) {
    showToast('Save failed: ' + e.message, 'error');
  }
}
</script>
</body>
</html>
"""


@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard():
    """
    Serve the visual configuration dashboard.

    The page itself is unprotected; sensitive operations (reading/writing
    config) require the user to supply the API key via the login form.
    """
    return HTMLResponse(content=_ADMIN_HTML)


# ===================================================================
# Global exception handlers
# ===================================================================


@app.exception_handler(HTTPException)
async def http_exception_handler(_request: Request, exc: HTTPException) -> JSONResponse:
    """Return consistent JSON for all HTTP exceptions."""
    return JSONResponse(
        status_code=exc.status_code,
        content={"status": "error", "detail": exc.detail},
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(_request: Request, exc: Exception) -> JSONResponse:
    """Catch-all for unexpected errors (log in production)."""
    return JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={
            "status": "error",
            "detail": "An unexpected internal error occurred.",
        },
    )


# ===================================================================
# Entry-point
# ===================================================================

# ---------------------------------------------------------------------------
# Auto-registration helper
# ---------------------------------------------------------------------------

def _auto_register(central_url: str, bridge_name: str = "", location: str = "", office: str = "", is_main: bool = False) -> bool:
    """
    Attempt to register this bridge with the central web app.
    Returns True on success, False on failure.

    If bridge_name is not provided, uses the hostname.
    If location is not provided, tries BRIDGE_LOCATION env var, then falls back
    to the bridge name.
    is_main — when True, marks this bridge as the main bridge for its location.
    """
    central_url = central_url.rstrip("/")

    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"

    if not bridge_name:
        bridge_name = os.getenv("BRIDGE_NAME", hostname)
    if not location:
        location = os.getenv("BRIDGE_LOCATION", "")
    if not location:
        location = bridge_name
    if not office:
        office = os.getenv("BRIDGE_OFFICE", "")

    # Use REGISTRATION_KEY from env if set
    reg_key = os.getenv("REGISTRATION_KEY", "register-me")

    # Use the permanent bridge UUID — stays the same even if IP/name changes
    bridge_id = settings.BRIDGE_ID

    payload = {
        "id": bridge_id,
        "name": bridge_name,
        "location": location,
        "office": office,
        "url": f"http://{local_ip}:{settings.PORT}",
        "api_key": settings.API_KEY,
        "fernet_key": settings.FERNET_KEY,
        "is_main": is_main,
    }

    try:
        print(f"\n  📡 Registering with central app at {central_url} ...")
        resp = _requests.post(
            f"{central_url}/api/register-bridge",
            json=payload,
            headers={"X-Registration-Key": reg_key},
            timeout=10,
        )
        if resp.status_code in (200, 201):
            data = resp.json()
            print(f"  ✅ {data.get('message', 'Registered successfully!')}")
            return True
        else:
            print(f"  ⚠️  Registration failed (HTTP {resp.status_code}): {resp.text[:200]}")
            return False
    except _requests.exceptions.ConnectionError:
        print(f"  ❌ Cannot reach central app at {central_url}")
        return False
    except Exception as exc:
        print(f"  ❌ Registration error: {exc}")
        return False


# ===================================================================
# Entry-point
# ===================================================================

if __name__ == "__main__":
    import uvicorn

    # --- CLI ---
    parser = argparse.ArgumentParser(description="Local Bridge Service")
    parser.add_argument(
        "--register",
        metavar="URL",
        nargs="?",
        const=os.getenv("CENTRAL_APP_URL", ""),
        default=None,
        help="Auto-register with central web app. Use env CENTRAL_APP_URL if no URL given.",
    )
    parser.add_argument(
        "--name",
        metavar="NAME",
        default=os.getenv("BRIDGE_NAME", ""),
        help="Friendly name for this bridge (e.g. 'Colombo Office'). Also read from BRIDGE_NAME env var.",
    )
    parser.add_argument(
        "--location",
        metavar="LOCATION",
        default=os.getenv("BRIDGE_LOCATION", ""),
        help="Human-readable location label (e.g. 'Colombo, Sri Lanka'). Also read from BRIDGE_LOCATION env var.",
    )
    parser.add_argument(
        "--office",
        metavar="OFFICE",
        default=os.getenv("BRIDGE_OFFICE", ""),
        help="Office name within the location (e.g. 'Head Office'). Also read from BRIDGE_OFFICE env var.",
    )
    parser.add_argument(
        "--main",
        action="store_true",
        default=False,
        help="Mark this bridge as the main bridge for its location.",
    )
    args = parser.parse_args()

    # --- Print startup banner ---
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"

    print("=" * 60)
    print("  🔌 Local Bridge Service")
    print("=" * 60)
    print(f"  Machine    : {hostname}")
    print(f"  Local IP   : {local_ip}")
    print(f"  OS         : {platform.system()} {platform.release()}")
    print(f"  Bridge URL : http://{local_ip}:{settings.PORT}")
    print(f"  Health     : http://{local_ip}:{settings.PORT}/health")
    print(f"  Info       : http://{local_ip}:{settings.PORT}/info")
    print(f"  Admin      : http://{local_ip}:{settings.PORT}/admin")
    print(f"  Data File  : {settings.EXCEL_FILE_PATH}")
    print("-" * 60)
    print("  💡 To make this bridge reachable from the internet:")
    print("     1. Port forwarding: forward port 8000 on your router")
    print("     2. Cloud tunnel:    ngrok http 8000")
    print("     3. Cloudflare Tunnel: cloudflared tunnel --url http://localhost:8000")
    print("=" * 60)

    # --- Auto-register if requested ---
    if args.register:
        _auto_register(args.register, args.name, args.location, args.office, args.main)
        print()

    uvicorn.run(
        app,
        host=settings.HOST,
        port=settings.PORT,
        reload=False,  # Set to True for development only
    )
